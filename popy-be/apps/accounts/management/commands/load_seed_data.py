"""Load example POS data from separate xlsx files in popy-be/seed-data/."""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_datetime

from apps.accounts.models import User
from apps.catalog.models import Category, Product
from apps.core.choices import Role, StockTransactionType
from apps.core.services import adjust_stock
from apps.inventory.models import StockAdjustment, StockTransaction
from apps.parties.models import Customer, Supplier
from apps.purchasing.models import Purchase, PurchaseItem
from apps.returns.models import PurchaseReturn, SalesReturn
from apps.sales.models import Sale, SaleItem
from apps.settings.models import NotificationSettings
from apps.shops.models import Shop

DEFAULT_DIR = Path(settings.BASE_DIR) / "seed-data"
SEED_FILES = (
    "shops",
    "users",
    "notification_settings",
    "categories",
    "products",
    "customers",
    "suppliers",
    "purchases",
    "purchase_items",
    "sales",
    "sale_items",
    "stock_adjustments",
    "sales_returns",
    "purchase_returns",
)


def _cell(row: dict, key: str, default=""):
    value = row.get(key, default)
    if value is None:
        return default
    return value


def _str(row: dict, key: str) -> str:
    value = _cell(row, key, "")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _bool(row: dict, key: str, default: bool = True) -> bool:
    value = _cell(row, key, default)
    if isinstance(value, bool):
        return value
    text = str(value).strip().upper()
    if text in {"", "NONE"}:
        return default
    return text in {"1", "TRUE", "YES", "Y"}


def _int(row: dict, key: str, default: int = 0) -> int:
    value = _cell(row, key, default)
    if value in ("", None):
        return default
    return int(value)


def _dec(row: dict, key: str, default: str = "0") -> Decimal:
    value = _cell(row, key, default)
    if value in ("", None):
        return Decimal(default)
    return Decimal(str(value))


def _dt(row: dict, key: str):
    value = _cell(row, key, None)
    if not value:
        return None
    if isinstance(value, datetime):
        when = value
    else:
        text = str(value).strip()
        when = parse_datetime(text.replace(" ", "T"))
        if when is None:
            try:
                when = datetime.fromisoformat(text)
            except ValueError:
                when = datetime.strptime(text, "%Y-%m-%d %H:%M:%S")
    if timezone.is_naive(when):
        when = timezone.make_aware(when, timezone.get_current_timezone())
    return when


def _set_timestamps(model, pk, when):
    if when is None:
        return
    updates = {"created_at": when}
    field_names = {f.name for f in model._meta.fields}
    if "updated_at" in field_names:
        updates["updated_at"] = when
    model.objects.filter(pk=pk).update(**updates)


def _parse_return_items(raw: str) -> list[dict]:
    text = (raw or "").strip()
    if not text:
        return []
    items = []
    for part in text.split("|"):
        sku, qty = part.split(":", 1)
        items.append({"product_id": sku.strip(), "quantity": int(qty)})
    return items


def read_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for raw in rows[1:]:
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        data.append({headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers)) if headers[i]})
    return data


def read_seed_dir(seed_dir: Path) -> dict[str, list[dict]]:
    missing = [name for name in SEED_FILES if not (seed_dir / f"{name}.xlsx").is_file()]
    if missing:
        raise CommandError(
            f"Missing seed files in {seed_dir}: " + ", ".join(f"{name}.xlsx" for name in missing)
        )
    return {name: read_xlsx(seed_dir / f"{name}.xlsx") for name in SEED_FILES}


class Command(BaseCommand):
    help = "Load example POS data from separate xlsx files in seed-data/"

    def add_arguments(self, parser):
        parser.add_argument(
            "--dir",
            dest="seed_dir",
            default=str(DEFAULT_DIR),
            help="Folder containing shops.xlsx, products.xlsx, sales.xlsx, ...",
        )
        parser.add_argument(
            "--reload",
            action="store_true",
            help="Delete previously loaded 2026 example transactions and replay stock from the xlsx files",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl  # noqa: F401
        except ImportError as exc:
            raise CommandError("openpyxl is required. Install with: pip install openpyxl") from exc

        seed_dir = Path(options["seed_dir"])
        if not seed_dir.is_dir():
            raise CommandError(f"Seed folder not found: {seed_dir}")

        sheets = read_seed_dir(seed_dir)

        with transaction.atomic():
            shops = self._load_shops(sheets["shops"])
            self._load_users(sheets["users"], shops)
            self._disable_notifications(shops)
            categories = self._load_categories(sheets["categories"], shops)
            products = self._load_products(sheets["products"], shops, categories)
            self._load_customers(sheets["customers"], shops)
            self._load_suppliers(sheets["suppliers"], shops)

            already_loaded = Sale.objects.filter(reference="SL-2026-0001").exists()
            if already_loaded and not options["reload"]:
                self._load_notification_settings(sheets["notification_settings"], shops)
                self.stdout.write(
                    self.style.WARNING(
                        "Example sales already exist. Master data was updated; "
                        "transactions were left unchanged. Re-run with --reload to replay them."
                    )
                )
                return

            if options["reload"]:
                self._purge_example_transactions(shops, products)

            self._apply_opening_stock(sheets["products"], products)
            self._load_purchases(sheets["purchases"], sheets["purchase_items"], shops, products)
            self._load_sales(sheets["sales"], sheets["sale_items"], shops, products)
            self._load_adjustments(sheets["stock_adjustments"], shops, products)
            self._load_sales_returns(sheets["sales_returns"], shops, products)
            self._load_purchase_returns(sheets["purchase_returns"], shops, products)
            self._load_notification_settings(sheets["notification_settings"], shops)

        self.stdout.write(self.style.SUCCESS(f"Loaded example data from {seed_dir}"))

    def _load_shops(self, rows: list[dict]) -> dict[str, Shop]:
        shops = {}
        for row in rows:
            code = _str(row, "code")
            shop, created = Shop.objects.update_or_create(
                code=code,
                defaults={
                    "name": _str(row, "name"),
                    "address": _str(row, "address"),
                    "phone": _str(row, "phone"),
                    "email": _str(row, "email"),
                    "is_active": _bool(row, "is_active", True),
                },
            )
            shops[code] = shop
            self.stdout.write(("Created" if created else "Updated") + f" shop {code}")
        return shops

    def _load_users(self, rows: list[dict], shops: dict[str, Shop]) -> None:
        for row in rows:
            email = _str(row, "email").lower()
            shop = shops.get(_str(row, "shop_code"))
            role = _str(row, "role") or Role.CASHIER
            defaults = {
                "name": _str(row, "name"),
                "role": role,
                "shop": shop,
                "is_active": _bool(row, "is_active", True),
                "is_staff": _bool(row, "is_staff", role == Role.SUPER_ADMIN),
                "is_superuser": _bool(row, "is_superuser", role == Role.SUPER_ADMIN),
            }
            user, created = User.objects.get_or_create(email=email, defaults=defaults)
            password = _str(row, "password")
            if created:
                user.set_password(password)
                user.save()
                self.stdout.write(self.style.SUCCESS(f"Created user {email}"))
            else:
                for field, value in defaults.items():
                    setattr(user, field, value)
                if password:
                    user.set_password(password)
                user.save()
                self.stdout.write(f"Updated user {email}")

    def _disable_notifications(self, shops: dict[str, Shop]) -> None:
        for shop in shops.values():
            settings_obj = NotificationSettings.load(shop)
            NotificationSettings.objects.filter(pk=settings_obj.pk).update(
                pos_checkout_email_enabled=False,
                pos_checkout_sms_enabled=False,
                pos_checkout_cashier_email_enabled=False,
                pos_checkout_cashier_sms_enabled=False,
                low_inventory_email_enabled=False,
                low_inventory_sms_enabled=False,
                new_customer_email_enabled=False,
                new_customer_sms_enabled=False,
                new_user_email_enabled=False,
                new_user_sms_enabled=False,
            )

    def _load_categories(self, rows: list[dict], shops: dict[str, Shop]) -> dict[tuple[str, str], Category]:
        categories = {}
        for row in rows:
            shop = shops[_str(row, "shop_code")]
            name = _str(row, "name")
            category, _ = Category.objects.update_or_create(
                shop=shop,
                name=name,
                defaults={"description": _str(row, "description")},
            )
            categories[(shop.code, name)] = category
        return categories

    def _load_products(self, rows: list[dict], shops: dict[str, Shop], categories) -> dict[tuple[str, str], Product]:
        products = {}
        for row in rows:
            shop = shops[_str(row, "shop_code")]
            sku = _str(row, "sku")
            category = categories[(shop.code, _str(row, "category_name"))]
            barcode = _str(row, "barcode") or None
            product, _ = Product.objects.update_or_create(
                shop=shop,
                sku=sku,
                defaults={
                    "name": _str(row, "name"),
                    "barcode": barcode,
                    "category": category,
                    "brand": _str(row, "brand") or None,
                    "unit": _str(row, "unit"),
                    "cost_price": _dec(row, "cost_price"),
                    "selling_price": _dec(row, "selling_price"),
                    "reorder_level": _int(row, "reorder_level"),
                    "status": _bool(row, "status", True),
                },
            )
            products[(shop.code, sku)] = product
        return products

    def _load_customers(self, rows: list[dict], shops: dict[str, Shop]) -> None:
        for row in rows:
            shop = shops[_str(row, "shop_code")]
            Customer.objects.update_or_create(
                shop=shop,
                phone=_str(row, "phone"),
                defaults={
                    "name": _str(row, "name"),
                    "email": _str(row, "email") or None,
                    "address": _str(row, "address") or None,
                    "loyalty_points": _int(row, "loyalty_points"),
                },
            )

    def _load_suppliers(self, rows: list[dict], shops: dict[str, Shop]) -> None:
        for row in rows:
            shop = shops[_str(row, "shop_code")]
            Supplier.objects.update_or_create(
                shop=shop,
                phone=_str(row, "phone"),
                defaults={
                    "name": _str(row, "name"),
                    "company_name": _str(row, "company_name") or None,
                    "email": _str(row, "email") or None,
                    "address": _str(row, "address") or None,
                },
            )

    def _load_notification_settings(self, rows: list[dict], shops: dict[str, Shop]) -> None:
        for row in rows:
            shop = shops[_str(row, "shop_code")]
            obj = NotificationSettings.load(shop)
            for field in (
                "pos_checkout_email_enabled",
                "pos_checkout_sms_enabled",
                "pos_checkout_cashier_email_enabled",
                "pos_checkout_cashier_sms_enabled",
                "low_inventory_email_enabled",
                "low_inventory_sms_enabled",
                "new_customer_email_enabled",
                "new_customer_sms_enabled",
                "new_user_email_enabled",
                "new_user_sms_enabled",
            ):
                setattr(obj, field, _bool(row, field, True))
            obj.low_inventory_alert_phone = _str(row, "low_inventory_alert_phone")
            obj.save()

    def _purge_example_transactions(self, shops: dict[str, Shop], products) -> None:
        shop_ids = [shop.id for shop in shops.values()]
        SalesReturn.objects.filter(shop_id__in=shop_ids, reference__startswith="SR-2026-").delete()
        PurchaseReturn.objects.filter(shop_id__in=shop_ids, reference__startswith="PR-2026-").delete()
        Sale.objects.filter(shop_id__in=shop_ids, reference__startswith="SL-2026-").delete()
        Purchase.objects.filter(shop_id__in=shop_ids, reference__startswith="PO-2026-").delete()
        product_ids = [p.id for p in products.values()]
        StockAdjustment.objects.filter(product_id__in=product_ids).delete()
        StockTransaction.objects.filter(product_id__in=product_ids).delete()
        Product.objects.filter(id__in=product_ids).update(stock_quantity=0)
        for product in products.values():
            product.stock_quantity = 0
        self.stdout.write("Removed previously loaded example transactions and reset stock")

    def _apply_opening_stock(self, rows: list[dict], products) -> None:
        opening_at = timezone.make_aware(datetime(2026, 7, 1, 8, 0, 0), timezone.get_current_timezone())
        for row in rows:
            product = products[(_str(row, "shop_code"), _str(row, "sku"))]
            product.refresh_from_db()
            qty = _int(row, "opening_stock")
            if qty == 0:
                continue
            if product.stock_quantity:
                Product.objects.filter(pk=product.pk).update(stock_quantity=0)
                product.stock_quantity = 0
            adjust_stock(
                product,
                qty,
                StockTransactionType.ADJUSTMENT,
                note="Opening stock (seed)",
                reference_type="Seed",
            )
            tx = product.stock_transactions.order_by("-id").first()
            if tx:
                StockTransaction.objects.filter(pk=tx.pk).update(created_at=opening_at)
            product.refresh_from_db()

    def _load_purchases(self, purchases, items, shops, products) -> None:
        items_by_ref: dict[tuple[str, str], list[dict]] = {}
        for row in items:
            items_by_ref.setdefault((_str(row, "shop_code"), _str(row, "purchase_reference")), []).append(row)

        for row in purchases:
            shop = shops[_str(row, "shop_code")]
            reference = _str(row, "reference")
            if Purchase.objects.filter(shop=shop, reference=reference).exists():
                continue
            supplier = Supplier.objects.get(shop=shop, phone=_str(row, "supplier_phone"))
            status = _str(row, "status")
            purchase = Purchase.objects.create(
                shop=shop,
                reference=reference,
                supplier=supplier,
                status=status,
                note=_str(row, "note") or None,
                total=Decimal("0"),
            )
            total = Decimal("0")
            for item in items_by_ref.get((shop.code, reference), []):
                product = products[(shop.code, _str(item, "sku"))]
                qty = _int(item, "quantity")
                cost = _dec(item, "cost_price")
                line = PurchaseItem.objects.create(
                    purchase=purchase,
                    product=product,
                    quantity=qty,
                    cost_price=cost,
                    line_total=cost * qty,
                )
                total += line.line_total
                if status == "RECEIVED":
                    adjust_stock(
                        product,
                        qty,
                        StockTransactionType.PURCHASE,
                        note=f"Purchase {reference}",
                        reference_type="Purchase",
                        reference_id=purchase.id,
                    )
                    product.cost_price = cost
                    product.save(update_fields=["cost_price", "updated_at"])
                    tx = product.stock_transactions.order_by("-id").first()
                    _set_timestamps(StockTransaction, tx.pk, _dt(row, "created_at"))
            purchase.total = total
            purchase.save(update_fields=["total"])
            _set_timestamps(Purchase, purchase.id, _dt(row, "created_at"))

    def _load_sales(self, sales, items, shops, products) -> None:
        items_by_ref: dict[tuple[str, str], list[dict]] = {}
        for row in items:
            items_by_ref.setdefault((_str(row, "shop_code"), _str(row, "sale_reference")), []).append(row)

        for row in sales:
            shop = shops[_str(row, "shop_code")]
            reference = _str(row, "reference")
            if Sale.objects.filter(shop=shop, reference=reference).exists():
                continue
            cashier = User.objects.get(email__iexact=_str(row, "cashier_email"))
            phone = _str(row, "customer_phone")
            customer = Customer.objects.filter(shop=shop, phone=phone).first() if phone else None
            sale = Sale.objects.create(
                shop=shop,
                reference=reference,
                customer=customer,
                subtotal=_dec(row, "subtotal"),
                discount=_dec(row, "discount"),
                tax=_dec(row, "tax"),
                total=_dec(row, "total"),
                payment_method=_str(row, "payment_method"),
                amount_paid=_dec(row, "amount_paid"),
                cashier=cashier,
            )
            for item in items_by_ref.get((shop.code, reference), []):
                product = products[(shop.code, _str(item, "sku"))]
                qty = _int(item, "quantity")
                unit_price = _dec(item, "unit_price")
                SaleItem.objects.create(
                    sale=sale,
                    product=product,
                    quantity=qty,
                    unit_price=unit_price,
                    total=unit_price * qty,
                )
                adjust_stock(
                    product,
                    -qty,
                    StockTransactionType.SALE,
                    note=f"Sale {reference}",
                    reference_type="Sale",
                    reference_id=sale.id,
                )
                tx = product.stock_transactions.order_by("-id").first()
                _set_timestamps(StockTransaction, tx.pk, _dt(row, "created_at"))
            _set_timestamps(Sale, sale.id, _dt(row, "created_at"))

    def _load_adjustments(self, rows: list[dict], shops, products) -> None:
        for row in rows:
            shop = shops[_str(row, "shop_code")]
            product = products[(shop.code, _str(row, "sku"))]
            qty = _int(row, "quantity")
            user = User.objects.filter(email__iexact=_str(row, "user_email")).first()
            adjust_stock(
                product,
                qty,
                StockTransactionType.ADJUSTMENT,
                note=_str(row, "note"),
                reference_type="StockAdjustment",
            )
            adj = StockAdjustment.objects.create(
                product=product,
                adjustment_type=_str(row, "adjustment_type"),
                quantity=qty,
                note=_str(row, "note") or None,
                user=user,
            )
            when = _dt(row, "created_at")
            _set_timestamps(StockAdjustment, adj.id, when)
            tx = product.stock_transactions.order_by("-id").first()
            _set_timestamps(StockTransaction, tx.pk, when)

    def _resolve_return_lines(self, raw_items: str, fallback_lines, products, shop_code: str):
        parsed = _parse_return_items(raw_items)
        if not parsed:
            return fallback_lines
        resolved = []
        for item in parsed:
            product = products[(shop_code, item["product_id"])]
            resolved.append((product, item["quantity"]))
        return resolved

    def _load_sales_returns(self, rows: list[dict], shops, products) -> None:
        for row in rows:
            shop = shops[_str(row, "shop_code")]
            reference = _str(row, "reference")
            if SalesReturn.objects.filter(shop=shop, reference=reference).exists():
                continue
            sale = Sale.objects.prefetch_related("items__product").get(
                shop=shop, reference=_str(row, "sale_reference")
            )
            fallback = [(item.product, item.quantity) for item in sale.items.all()]
            lines = self._resolve_return_lines(_str(row, "items"), fallback, products, shop.code)
            payload_items = (
                [{"productId": product.id, "quantity": qty} for product, qty in lines]
                if _str(row, "items")
                else []
            )
            for product, qty in lines:
                adjust_stock(
                    product,
                    qty,
                    StockTransactionType.RETURN,
                    note=f"Sales return {reference}",
                    reference_type="SalesReturn",
                )
                tx = product.stock_transactions.order_by("-id").first()
                _set_timestamps(StockTransaction, tx.pk, _dt(row, "created_at"))
            ret = SalesReturn.objects.create(
                shop=shop,
                reference=reference,
                sale=sale,
                reason=_str(row, "reason"),
                refund_amount=_dec(row, "refund_amount"),
                items=payload_items,
            )
            _set_timestamps(SalesReturn, ret.id, _dt(row, "created_at"))

    def _load_purchase_returns(self, rows: list[dict], shops, products) -> None:
        for row in rows:
            shop = shops[_str(row, "shop_code")]
            reference = _str(row, "reference")
            if PurchaseReturn.objects.filter(shop=shop, reference=reference).exists():
                continue
            purchase = Purchase.objects.prefetch_related("items__product").get(
                shop=shop, reference=_str(row, "purchase_reference")
            )
            fallback = [(item.product, item.quantity) for item in purchase.items.all()]
            lines = self._resolve_return_lines(_str(row, "items"), fallback, products, shop.code)
            payload_items = (
                [{"productId": product.id, "quantity": qty} for product, qty in lines]
                if _str(row, "items")
                else []
            )
            for product, qty in lines:
                adjust_stock(
                    product,
                    -qty,
                    StockTransactionType.RETURN,
                    note=f"Purchase return {reference}",
                    reference_type="PurchaseReturn",
                )
                tx = product.stock_transactions.order_by("-id").first()
                _set_timestamps(StockTransaction, tx.pk, _dt(row, "created_at"))
            ret = PurchaseReturn.objects.create(
                shop=shop,
                reference=reference,
                purchase=purchase,
                reason=_str(row, "reason"),
                amount=_dec(row, "amount"),
                items=payload_items,
            )
            _set_timestamps(PurchaseReturn, ret.id, _dt(row, "created_at"))
