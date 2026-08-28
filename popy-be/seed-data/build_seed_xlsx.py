"""
Write one .xlsx file per table into this folder.

Run from popy-be:
    python seed-data/build_seed_xlsx.py

    cd popy-be
    python -m pip install -r requirements.txt
    python manage.py load_seed_data 
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TAX_RATE = Decimal("0.10")
YEAR = 2026

# ---------------------------------------------------------------------------
# Catalog
# ---------------------------------------------------------------------------

# (name, sku, barcode_stem, category, brand, unit, cost, sell, reorder,
#  opening_MAIN, opening_KDY or None to omit from Kandy)
PRODUCTS = [
    # Beverages
    ('Coca-Cola 1.5L', 'BEV-001', '479000100001', 'Beverages', 'Coca-Cola', 'bottle', '340', '420', 20, 90, 40),
    ('Sprite 1.5L', 'BEV-002', '479000100002', 'Beverages', 'Coca-Cola', 'bottle', '340', '420', 18, 70, 28),
    ('Elephant House Cream Soda 1L', 'BEV-003', '479000100003', 'Beverages', 'Elephant House', 'bottle', '170', '230', 15, 55, 20),
    ('Mineral Water 1L', 'BEV-004', '479000100004', 'Beverages', 'Aqua Safe', 'bottle', '60', '90', 40, 160, 70),
    ('Nestlé Milo 400g', 'BEV-005', '479000100005', 'Beverages', 'Nestlé', 'tin', '850', '1050', 10, 36, 14),
    ('Dilmah Tea 100 bags', 'BEV-006', '479000100006', 'Beverages', 'Dilmah', 'box', '580', '720', 12, 42, 16),
    ('House Coffee 250g', 'BEV-007', '479000100007', 'Beverages', 'House', 'pack', '480', '620', 8, 24, None),
    # Snacks
    ('Potato Chips 150g', 'SNK-001', '479000100011', 'Snacks', 'Munchee', 'pack', '210', '280', 25, 18, 8),
    ('Maliban Lemon Puff', 'SNK-002', '479000100012', 'Snacks', 'Maliban', 'pack', '160', '220', 20, 60, 24),
    ('Munchee Super Cream Cracker', 'SNK-003', '479000100013', 'Snacks', 'Munchee', 'pack', '170', '230', 15, 48, 18),
    ('Chocolate Digestive 200g', 'SNK-004', '479000100014', 'Snacks', 'Maliban', 'pack', '280', '380', 12, 32, None),
    ('MD Roasted Peanuts 200g', 'SNK-005', '479000100015', 'Snacks', 'MD', 'pack', '220', '300', 12, 26, 10),
    # Grocery
    ('Keeri Samba Rice 5kg', 'GRC-001', '479000100021', 'Grocery', 'Nipuna', 'bag', '2000', '2400', 8, 28, 12),
    ('White Rice 5kg', 'GRC-002', '479000100022', 'Grocery', 'Nipuna', 'bag', '1450', '1750', 8, 10, 3),
    ('Prima Wheat Flour 1kg', 'GRC-003', '479000100023', 'Grocery', 'Prima', 'pack', '240', '300', 15, 44, 16),
    ('White Sugar 1kg', 'GRC-004', '479000100024', 'Grocery', 'Harischandra', 'pack', '240', '280', 20, 70, 30),
    ('Cooking Oil 1L', 'GRC-005', '479000100025', 'Grocery', 'Fortune', 'bottle', '900', '1050', 12, 36, 14),
    ('Maggi Noodles 5-pack', 'GRC-006', '479000100026', 'Grocery', 'Nestlé', 'pack', '380', '480', 20, 80, 32),
    ('MD Tomato Ketchup 400g', 'GRC-007', '479000100027', 'Grocery', 'MD', 'bottle', '320', '400', 10, 30, None),
    ('Coconut Milk 400ml', 'GRC-008', '479000100028', 'Grocery', 'Maggi', 'tin', '180', '240', 18, 48, 18),
    ('Red Dhal 500g', 'GRC-009', '479000100029', 'Grocery', 'House', 'pack', '280', '350', 15, 40, 16),
    ('Iodised Salt 1kg', 'GRC-010', '479000100030', 'Grocery', 'House', 'pack', '70', '100', 20, 80, 36),
    # Dairy
    ('Anchor Milk Powder 400g', 'DRY-001', '479000100041', 'Dairy', 'Anchor', 'tin', '900', '1080', 10, 28, 12),
    ('Kotmale Fresh Milk 1L', 'DRY-002', '479000100042', 'Dairy', 'Kotmale', 'bottle', '320', '400', 12, 22, 8),
    ('Highland Yoghurt 80ml', 'DRY-003', '479000100043', 'Dairy', 'Highland', 'pcs', '55', '80', 30, 60, 24),
    ('Anchor Cheese Slices 200g', 'DRY-004', '479000100044', 'Dairy', 'Anchor', 'pack', '650', '800', 8, 16, None),
    ('Anchor Butter 227g', 'DRY-005', '479000100045', 'Dairy', 'Anchor', 'pack', '560', '700', 8, 14, 6),
    # Bakery
    ('White Bread Loaf', 'BKY-001', '479000100051', 'Bakery', 'Happy Cow', 'loaf', '110', '150', 15, 22, 10),
    ('Sweet Buns 6-pack', 'BKY-002', '479000100052', 'Bakery', 'Maliban', 'pack', '190', '260', 12, 24, 8),
    ('Chocolate Cake Slice', 'BKY-003', '479000100053', 'Bakery', 'House', 'pcs', '200', '280', 10, 2, 0),
    # Personal Care
    ('Signal Toothpaste 120g', 'PER-001', '479000100061', 'Personal Care', 'Unilever', 'pcs', '300', '380', 12, 34, 14),
    ('Lifebuoy Soap', 'PER-002', '479000100062', 'Personal Care', 'Unilever', 'pcs', '85', '115', 20, 50, 22),
    ('Sunsilk Shampoo 180ml', 'PER-003', '479000100063', 'Personal Care', 'Unilever', 'bottle', '400', '520', 10, 22, None),
    # Baby Care
    ('Baby Diapers M 22', 'BAB-001', '479000100071', 'Baby Care', 'Huggies', 'pack', '1400', '1750', 6, 16, 6),
    # Household
    ('Vim Dishwash 500ml', 'HLD-001', '479000100081', 'Household', 'Unilever', 'bottle', '280', '350', 12, 30, 12),
    ('Sunlight Detergent 1kg', 'HLD-002', '479000100082', 'Household', 'Unilever', 'pack', '420', '520', 10, 22, 8),
    ('Facial Tissue 4-pack', 'HLD-003', '479000100083', 'Household', 'Lotus', 'pack', '340', '430', 15, 36, 14),
    ('Garbage Bags 30s', 'HLD-004', '479000100084', 'Household', 'House', 'pack', '230', '300', 10, 26, None),
    # Frozen
    ('Chicken Sausages 500g', 'FRZ-001', '479000100091', 'Frozen', 'Keells', 'pack', '650', '800', 8, 18, 8),
    ('Vanilla Ice Cream 1L', 'FRZ-002', '479000100092', 'Frozen', 'Elephant House', 'tub', '800', '1020', 6, 12, 4),
    ('Frozen Green Peas 500g', 'FRZ-003', '479000100093', 'Frozen', 'Elephant House', 'pack', '350', '450', 8, 20, None),
    # Produce
    ('Bananas 1kg', 'PRD-001', '479000100101', 'Produce', 'Local', 'kg', '180', '250', 10, 16, 8),
    ('Tomatoes 1kg', 'PRD-002', '479000100102', 'Produce', 'Local', 'kg', '250', '350', 8, 11, 4),
    ('King Coconut', 'PRD-003', '479000100103', 'Produce', 'Local', 'pcs', '80', '120', 15, 28, 12),
    # Beverages
    ('Pepsi 1.5L', 'BEV-008', '479000100111', 'Beverages', 'Pepsi', 'bottle', '330', '400', 20, 60, 24),
    ('Fanta Orange 1.5L', 'BEV-009', '479000100112', 'Beverages', 'Coca-Cola', 'bottle', '330', '400', 18, 50, None),
    ('Elephant Ginger Beer 400ml', 'BEV-010', '479000100113', 'Beverages', 'Elephant House', 'bottle', '130', '180', 20, 40, 16),
    ('Necto Grape 1.5L', 'BEV-011', '479000100114', 'Beverages', 'Necto', 'bottle', '330', '400', 18, 45, None),
    ('KIST Mango Nectar 1L', 'BEV-012', '479000100115', 'Beverages', 'KIST', 'tetra', '220', '320', 15, 30, 12),
    ('Nescafé Classic Coffee 200g', 'BEV-013', '479000100116', 'Beverages', 'Nescafé', 'jar', '780', '980', 8, 20, None),
    ('Basilur Ceylon Tea 100 bags', 'BEV-014', '479000100117', 'Beverages', 'Basilur', 'box', '600', '780', 10, 24, None),
    ('Soda Water 400ml', 'BEV-015', '479000100118', 'Beverages', 'Elephant House', 'bottle', '60', '100', 25, 60, 24),
    # Snacks
    ('Ritzbury Milk Chocolate 80g', 'SNK-006', '479000100121', 'Snacks', 'Ritzbury', 'bar', '220', '320', 15, 30, 12),
    ('KitKat 4 Finger', 'SNK-007', '479000100122', 'Snacks', 'Nestlé', 'bar', '160', '220', 20, 36, None),
    ('Munchee Marie Biscuits', 'SNK-008', '479000100123', 'Snacks', 'Munchee', 'pack', '140', '190', 20, 40, 16),
    ('Uswatte Cashews 100g', 'SNK-009', '479000100124', 'Snacks', 'Uswatte', 'pack', '350', '480', 10, 18, None),
    ('Popcorn Butter 90g', 'SNK-010', '479000100125', 'Snacks', 'House', 'pack', '130', '190', 15, 24, None),
    # Grocery
    ('MD Chilli Powder 100g', 'GRC-011', '479000100131', 'Grocery', 'MD', 'pack', '180', '250', 15, 30, 12),
    ('MD Curry Powder 100g', 'GRC-012', '479000100132', 'Grocery', 'MD', 'pack', '200', '270', 15, 28, None),
    ('Turmeric Powder 100g', 'GRC-013', '479000100133', 'Grocery', 'MD', 'pack', '90', '130', 20, 32, None),
    ('Rani Semolina 1kg', 'GRC-014', '479000100134', 'Grocery', 'Rani', 'pack', '220', '300', 12, 20, None),
    ('Atta Flour 1kg', 'GRC-015', '479000100135', 'Grocery', 'Prima', 'pack', '190', '260', 15, 26, 10),
    ('Ran Kew Coconut Oil 750ml', 'GRC-016', '479000100136', 'Grocery', 'Ran Kew', 'bottle', '780', '980', 10, 18, None),
    ('New Zealand Tuna Canned 185g', 'GRC-017', '479000100137', 'Grocery', 'New Zealand', 'tin', '320', '420', 12, 30, 12),
    ('Rich Life Sardines 425g', 'GRC-018', '479000100138', 'Grocery', 'Rich Life', 'tin', '380', '480', 10, 22, None),
    ('Delmege Strawberry Jam 500g', 'GRC-019', '479000100139', 'Grocery', 'Delmege', 'jar', '480', '620', 8, 14, None),
    ('Raigam Soya Meat 90g', 'GRC-020', '479000100140', 'Grocery', 'Raigam', 'pack', '140', '190', 15, 26, None),
    # Dairy
    ('Highland Condensed Milk 397g', 'DRY-006', '479000100141', 'Dairy', 'Highland', 'tin', '320', '420', 12, 24, 10),
    ('Nespray Milk Powder 400g', 'DRY-007', '479000100142', 'Dairy', 'Nespray', 'tin', '950', '1150', 10, 20, None),
    ('Eggs (tray of 10)', 'DRY-008', '479000100143', 'Dairy', 'Local', 'tray', '320', '400', 15, 30, 14),
    ('Elephant House Curd 500g', 'DRY-009', '479000100144', 'Dairy', 'Elephant House', 'pot', '280', '380', 10, 18, None),
    # Bakery
    ('Wonder Bake Fruit Cake 200g', 'BKY-004', '479000100151', 'Bakery', 'Wonder Bake', 'pack', '300', '420', 10, 16, None),
    ('Sandwich Bread Loaf', 'BKY-005', '479000100152', 'Bakery', 'Happy Cow', 'loaf', '130', '180', 15, 20, 8),
    ('Egg Rolls 4-pack', 'BKY-006', '479000100153', 'Bakery', 'House', 'pack', '200', '280', 10, 14, None),
    # Personal Care
    ('Colgate Toothpaste 100g', 'PER-004', '479000100161', 'Personal Care', 'Colgate', 'pcs', '280', '370', 15, 28, 12),
    ('Dettol Antiseptic Liquid 125ml', 'PER-005', '479000100162', 'Personal Care', 'Dettol', 'bottle', '320', '420', 10, 18, None),
    ('Head & Shoulders Shampoo 180ml', 'PER-006', '479000100163', 'Personal Care', 'P&G', 'bottle', '480', '620', 10, 16, None),
    ('Nivea Body Lotion 200ml', 'PER-007', '479000100164', 'Personal Care', 'Nivea', 'bottle', '650', '850', 8, 12, None),
    # Baby Care
    ("Johnson's Baby Powder 200g", 'BAB-002', '479000100171', 'Baby Care', "Johnson's", 'pcs', '480', '620', 8, 14, 6),
    ('Baby Wipes 80s', 'BAB-003', '479000100172', 'Baby Care', 'Huggies', 'pack', '380', '520', 10, 18, None),
    # Household
    ('Comfort Fabric Softener 750ml', 'HLD-005', '479000100181', 'Household', 'Unilever', 'bottle', '480', '620', 10, 16, None),
    ('Rose Toilet Tissue 10-roll', 'HLD-006', '479000100182', 'Household', 'Rose', 'pack', '780', '980', 8, 14, 6),
    ('Harpic Toilet Cleaner 500ml', 'HLD-007', '479000100183', 'Household', 'Harpic', 'bottle', '380', '480', 12, 20, None),
    ('Mortein Insect Spray 300ml', 'HLD-008', '479000100184', 'Household', 'Mortein', 'can', '480', '620', 8, 12, None),
    # Frozen
    ('Frozen Chicken Wings 1kg', 'FRZ-004', '479000100191', 'Frozen', 'Keells', 'pack', '950', '1200', 6, 14, 6),
    ('Frozen Fish Fillets 500g', 'FRZ-005', '479000100192', 'Frozen', 'Keells', 'pack', '650', '850', 8, 12, None),
    # Produce
    ('Carrots 1kg', 'PRD-004', '479000100201', 'Produce', 'Local', 'kg', '180', '280', 10, 16, 6),
    ('Cabbage 1kg', 'PRD-005', '479000100202', 'Produce', 'Local', 'kg', '130', '200', 10, 14, None),
    ('Red Onions 1kg', 'PRD-006', '479000100203', 'Produce', 'Local', 'kg', '280', '380', 12, 20, 8),
    ('Potatoes 1kg', 'PRD-007', '479000100204', 'Produce', 'Local', 'kg', '220', '320', 12, 18, None),
    ('Green Chilies 250g', 'PRD-008', '479000100205', 'Produce', 'Local', 'pack', '100', '160', 15, 20, None),
    # Meat & Poultry
    ('Chicken Whole 1kg', 'MEAT-001', '479000100211', 'Meat & Poultry', 'Local', 'kg', '850', '1050', 8, 10, 4),
    ('Chicken Breast 1kg', 'MEAT-002', '479000100212', 'Meat & Poultry', 'Local', 'kg', '1050', '1300', 8, 8, None),
]
CATEGORIES = [
    ("Beverages", "Soft drinks, water, tea, and coffee"),
    ("Snacks", "Chips, biscuits, and packaged snacks"),
    ("Grocery", "Rice, flour, oil, and dry goods"),
    ("Dairy", "Milk, yoghurt, cheese, and butter"),
    ("Bakery", "Bread, buns, and cakes"),
    ("Personal Care", "Soap, toothpaste, and shampoo"),
    ("Baby Care", "Nappies and baby essentials"),
    ("Household", "Cleaning and home supplies"),
    ("Frozen", "Frozen meat, vegetables, and ice cream"),
    ("Produce", "Fresh fruit, vegetables, and coconuts"),
    ("Meat & Poultry", "Fresh chicken and poultry"),
]

SHOPS = [
    {
        "code": "MAIN",
        "name": "Popy Super Market",
        "address": "42 Galle Road, Colombo 03",
        "phone": "0112345678",
        "email": "colombo@popy.lk",
        "is_active": True,
    },
    {
        "code": "KDY",
        "name": "Popy Kandy Branch",
        "address": "18 Peradeniya Road, Kandy",
        "phone": "0812223344",
        "email": "kandy@popy.lk",
        "is_active": True,
    },
]

USERS = [
    {
        "email": "kavindu@gmail.com",
        "password": "12!@qwas",
        "name": "Kavindu Perera",
        "role": "SUPER_ADMIN",
        "shop_code": "MAIN",
        "is_active": True,
        "is_staff": True,
        "is_superuser": True,
    },
    {
        "email": "admin@test.com",
        "password": "123456",
        "name": "Admin User",
        "role": "SUPER_ADMIN",
        "shop_code": "MAIN",
        "is_active": True,
        "is_staff": True,
        "is_superuser": True,
    },
    {
        "email": "nadee@gmail.com",
        "password": "Ca!@1234",
        "name": "Nadeesha Fernando",
        "role": "MANAGER",
        "shop_code": "MAIN",
        "is_active": True,
        "is_staff": False,
        "is_superuser": False,
    },
    {
        "email": "cashier@gmail.com",
        "password": "ca@12345",
        "name": "Cashier Silva",
        "role": "CASHIER",
        "shop_code": "MAIN",
        "is_active": True,
        "is_staff": False,
        "is_superuser": False,
    },
    {
        "email": "sudubole@gmail.com",
        "password": "User@123",
        "name": "Sudu Bole",
        "role": "INVENTORY_OFFICER",
        "shop_code": "MAIN",
        "is_active": True,
        "is_staff": False,
        "is_superuser": False,
    },
    {
        "email": "user@gmail.com",
        "password": "user12345",
        "name": "Kandy Cashier",
        "role": "CASHIER",
        "shop_code": "KDY",
        "is_active": True,
        "is_staff": False,
        "is_superuser": False,
    },
    {
        "email": "kandy.manager@popy.lk",
        "password": "Manager@123",
        "name": "Amali Jayasuriya",
        "role": "MANAGER",
        "shop_code": "KDY",
        "is_active": True,
        "is_staff": False,
        "is_superuser": False,
    },
    {
        "email": "kandy.stock@popy.lk",
        "password": "Stock@123",
        "name": "Ruwan Bandara",
        "role": "INVENTORY_OFFICER",
        "shop_code": "KDY",
        "is_active": True,
        "is_staff": False,
        "is_superuser": False,
    },
]

CUSTOMERS = {
    "MAIN": [
        ("Walk-in Customer", "0000000000", None, "Counter sales", 0),
        ("Jane Fernando", "0771234501", "jane.fernando@example.com", "12 Flower Road, Colombo 07", 120),
        ("John Perera", "0771234502", "john.perera@example.com", "88 Bauddhaloka Mawatha, Colombo 04", 80),
        ("Ayesha Silva", "0771234503", "ayesha.silva@example.com", "5 Park Street, Colombo 02", 45),
        ("Nuwan Jayawardena", "0771234504", "nuwan.j@example.com", "21 Marine Drive, Mount Lavinia", 210),
        ("Dilani Wickramasinghe", "0771234505", "dilani.w@example.com", "9 Ward Place, Colombo 07", 15),
        ("Kasun Rathnayake", "0771234506", "kasun.r@example.com", "34 High Level Road, Nugegoda", 60),
        ("Fathima Rizwan", "0771234507", "fathima.r@example.com", "16 Messenger Street, Colombo 12", 0),
        ("Priyanka Mendis", "0771234508", "priyanka.m@example.com", "70 Duplication Road, Colombo 03", 95),
        ("Tharindu Gunasekara", "0771234509", "tharindu.g@example.com", "3 Templers Road, Mount Lavinia", 30),
        ("Sanduni Amarasinghe", "0771234510", None, "14 Havelock Road, Colombo 05", 0),
        ("Mohamed Imran", "0771234511", "imran.m@example.com", "22 Galle Face Court, Colombo 03", 150),
    ],
    "KDY": [
        ("Walk-in Customer", "0000000000", None, "Counter sales", 0),
        ("Chamari Bandara", "0752223301", "chamari.b@example.com", "11 Cross Street, Kandy", 70),
        ("Lakmal Herath", "0752223302", "lakmal.h@example.com", "8 DS Senanayake Veediya, Kandy", 40),
        ("Nirosha Kumari", "0752223303", "nirosha.k@example.com", "27 Peradeniya Road, Kandy", 25),
        ("Asela Dissanayake", "0752223304", None, "4 Katugastota Road, Kandy", 0),
        ("Ishara Weerasinghe", "0752223305", "ishara.w@example.com", "19 Ampitiya Road, Kandy", 110),
    ],
}

SUPPLIERS = {
    "MAIN": [
        ("Rohan de Silva", "Ceylon Beverages Ltd", "0112555001", "sales@ceylonbev.lk", "Industrial Zone, Biyagama"),
        ("Nimal Perera", "Prima Ceylon (Pvt) Ltd", "0112555002", "orders@prima.lk", "Ja-Ela"),
        ("Shanika Fonseka", "Fonterra Brands Lanka", "0112555003", "trade@fonterra.lk", "Biyagama"),
        ("Harsha Alwis", "Unilever Sri Lanka", "0112555004", "orders@unilever.lk", "Horana"),
        ("Malini Jayasuriya", "Maliban Biscuit Manufactories", "0112555005", "trade@maliban.lk", "Ratmalana"),
        ("Saman Kumara", "Cargills Food City Wholesale", "0112555006", "wholesale@cargills.lk", "Grandpass, Colombo"),
        ("Gayan Wickrama", "Keells Frozen Foods", "0112555007", "frozen@keells.lk", "Mulleriyawa"),
        ("Sunil Farmer Co-op", "Hill Country Produce", "0112555008", "produce@hillcountry.lk", "Nuwara Eliya"),
    ],
    "KDY": [
        ("Ruwan Ekanayake", "Kandy Cash & Carry", "0812200101", "sales@kandycnc.lk", "Katugastota"),
        ("Thilini Abey", "Central Distributors", "0812200102", "orders@centraldist.lk", "Kandy"),
        ("Mahinda Silva", "Upcountry Dairy Hub", "0812200103", "dairy@upcountry.lk", "Peradeniya"),
        ("Kamal Rajapaksha", "Hill Country Produce", "0812200104", "kandy@hillcountry.lk", "Gampola"),
    ],
}


def money(value) -> Decimal:
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def compute_tax(subtotal: Decimal, discount: Decimal) -> Decimal:
    taxable = subtotal - discount
    if taxable < 0:
        taxable = Decimal("0.00")
    return (taxable * TAX_RATE).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def dt(year: int, month: int, day: int, hour: int, minute: int = 0) -> datetime:
    return datetime(year, month, day, hour, minute, 0)


class SeedBuilder:
    def __init__(self) -> None:
        self.products: dict[tuple[str, str], dict] = {}
        self.stock: dict[tuple[str, str], int] = {}
        self.product_rows: list[dict] = []
        self.category_rows: list[dict] = []
        self.purchase_rows: list[dict] = []
        self.purchase_item_rows: list[dict] = []
        self.sale_rows: list[dict] = []
        self.sale_item_rows: list[dict] = []
        self.adjustment_rows: list[dict] = []
        self.sales_return_rows: list[dict] = []
        self.purchase_return_rows: list[dict] = []
        self._sale_seq: dict[str, int] = {"MAIN": 0, "KDY": 0}
        self._purchase_seq: dict[str, int] = {"MAIN": 0, "KDY": 0}
        self._sr_seq: dict[str, int] = {"MAIN": 0, "KDY": 0}
        self._pr_seq: dict[str, int] = {"MAIN": 0, "KDY": 0}

    def next_ref(self, prefix: str, shop: str, seq_map: dict[str, int]) -> str:
        seq_map[shop] += 1
        return f"{prefix}-{YEAR}-{seq_map[shop]:04d}"

    def apply_stock(self, shop: str, sku: str, delta: int, reason: str) -> None:
        key = (shop, sku)
        if key not in self.stock:
            raise ValueError(f"Unknown product {sku} in {shop} ({reason})")
        new_qty = self.stock[key] + delta
        if new_qty < 0:
            raise ValueError(
                f"Negative stock for {sku} in {shop}: {self.stock[key]}{delta:+d} after {reason}"
            )
        self.stock[key] = new_qty

    def build_catalog(self) -> None:
        for shop in SHOPS:
            for name, description in CATEGORIES:
                self.category_rows.append(
                    {"shop_code": shop["code"], "name": name, "description": description}
                )

        for (
            name,
            sku,
            barcode_stem,
            category,
            brand,
            unit,
            cost,
            sell,
            reorder,
            opening_main,
            opening_kdy,
        ) in PRODUCTS:
            openings = {"MAIN": opening_main, "KDY": opening_kdy}
            for shop_code, opening in openings.items():
                if opening is None:
                    continue
                barcode = barcode_stem if shop_code == "MAIN" else barcode_stem.replace("4790001", "4790002", 1)
                row = {
                    "shop_code": shop_code,
                    "name": name,
                    "sku": sku,
                    "barcode": barcode,
                    "category_name": category,
                    "brand": brand,
                    "unit": unit,
                    "cost_price": money(cost*130),
                    "selling_price": money(sell*130),
                    "reorder_level": reorder,
                    "opening_stock": opening,
                    "status": True,
                }
                self.product_rows.append(row)
                self.products[(shop_code, sku)] = row
                self.stock[(shop_code, sku)] = 0
                self.apply_stock(shop_code, sku, opening, "opening stock")

    def add_purchase(
        self,
        shop: str,
        supplier_phone: str,
        status: str,
        created_at: datetime,
        lines: list[tuple[str, int]],
        note: str = "",
    ) -> str:
        reference = self.next_ref("PO", shop, self._purchase_seq)
        self.purchase_rows.append(
            {
                "shop_code": shop,
                "reference": reference,
                "supplier_phone": supplier_phone,
                "status": status,
                "note": note,
                "created_at": created_at,
            }
        )
        for sku, qty in lines:
            product = self.products[(shop, sku)]
            self.purchase_item_rows.append(
                {
                    "shop_code": shop,
                    "purchase_reference": reference,
                    "sku": sku,
                    "quantity": qty,
                    "cost_price": product["cost_price"],
                }
            )
            if status == "RECEIVED":
                self.apply_stock(shop, sku, qty, f"purchase {reference}")
        return reference

    def add_sale(
        self,
        shop: str,
        created_at: datetime,
        cashier_email: str,
        customer_phone: str | None,
        lines: list[tuple[str, int]],
        payment_method: str = "CASH",
        discount: Decimal | int | str = 0,
        overpay: Decimal | int | str | None = None,
    ) -> str:
        reference = self.next_ref("SL", shop, self._sale_seq)
        discount = money(discount)
        subtotal = Decimal("0.00")
        item_rows = []
        for sku, qty in lines:
            product = self.products[(shop, sku)]
            unit_price = product["selling_price"]
            line_total = money(unit_price * qty)
            subtotal += line_total
            item_rows.append(
                {
                    "shop_code": shop,
                    "sale_reference": reference,
                    "sku": sku,
                    "quantity": qty,
                    "unit_price": unit_price,
                }
            )
            self.apply_stock(shop, sku, -qty, f"sale {reference}")

        tax = compute_tax(subtotal, discount)
        total = money(subtotal - discount + tax)
        amount_paid = total if overpay is None else money(total + money(overpay))

        self.sale_rows.append(
            {
                "shop_code": shop,
                "reference": reference,
                "customer_phone": customer_phone or "",
                "cashier_email": cashier_email,
                "payment_method": payment_method,
                "discount": discount,
                "tax": tax,
                "subtotal": subtotal,
                "total": total,
                "amount_paid": amount_paid,
                "created_at": created_at,
            }
        )
        self.sale_item_rows.extend(item_rows)
        return reference

    def add_adjustment(
        self,
        shop: str,
        sku: str,
        adjustment_type: str,
        quantity: int,
        note: str,
        user_email: str,
        created_at: datetime,
    ) -> None:
        self.apply_stock(shop, sku, quantity, f"adjustment {sku}")
        self.adjustment_rows.append(
            {
                "shop_code": shop,
                "sku": sku,
                "adjustment_type": adjustment_type,
                "quantity": quantity,
                "note": note,
                "user_email": user_email,
                "created_at": created_at,
            }
        )

    def add_sales_return(
        self,
        shop: str,
        sale_reference: str,
        reason: str,
        created_at: datetime,
        items: list[tuple[str, int]] | None = None,
    ) -> None:
        sale = next(row for row in self.sale_rows if row["reference"] == sale_reference and row["shop_code"] == shop)
        sale_items = [
            row for row in self.sale_item_rows if row["sale_reference"] == sale_reference and row["shop_code"] == shop
        ]
        if items is None:
            restore = [(row["sku"], row["quantity"]) for row in sale_items]
            items_cell = ""
            refund = sale["total"]
        else:
            restore = items
            items_cell = "|".join(f"{sku}:{qty}" for sku, qty in items)
            refund = Decimal("0.00")
            for sku, qty in items:
                unit = next(row["unit_price"] for row in sale_items if row["sku"] == sku)
                refund += money(unit * qty)
            refund = money(refund + compute_tax(refund, Decimal("0")))

        for sku, qty in restore:
            self.apply_stock(shop, sku, qty, f"sales return {sale_reference}")

        self.sales_return_rows.append(
            {
                "shop_code": shop,
                "reference": self.next_ref("SR", shop, self._sr_seq),
                "sale_reference": sale_reference,
                "reason": reason,
                "refund_amount": refund,
                "items": items_cell,
                "created_at": created_at,
            }
        )

    def add_purchase_return(
        self,
        shop: str,
        purchase_reference: str,
        reason: str,
        created_at: datetime,
        items: list[tuple[str, int]] | None = None,
    ) -> None:
        purchase_items = [
            row
            for row in self.purchase_item_rows
            if row["purchase_reference"] == purchase_reference and row["shop_code"] == shop
        ]
        if items is None:
            restore = [(row["sku"], row["quantity"]) for row in purchase_items]
            items_cell = ""
        else:
            restore = items
            items_cell = "|".join(f"{sku}:{qty}" for sku, qty in items)

        amount = Decimal("0.00")
        for sku, qty in restore:
            cost = next(row["cost_price"] for row in purchase_items if row["sku"] == sku)
            amount += money(cost * qty)
            self.apply_stock(shop, sku, -qty, f"purchase return {purchase_reference}")

        self.purchase_return_rows.append(
            {
                "shop_code": shop,
                "reference": self.next_ref("PR", shop, self._pr_seq),
                "purchase_reference": purchase_reference,
                "reason": reason,
                "amount": money(amount),
                "items": items_cell,
                "created_at": created_at,
            }
        )

    def build_purchases(self) -> None:
        self.add_purchase(
            "MAIN",
            "0112555001",
            "RECEIVED",
            dt(2026, 7, 18, 10, 15),
            [("BEV-001", 48), ("BEV-002", 36), ("BEV-003", 24), ("BEV-004", 96)],
            "Monthly soft-drink restock",
        )
        self.add_purchase(
            "MAIN",
            "0112555002",
            "RECEIVED",
            dt(2026, 7, 20, 9, 40),
            [("GRC-001", 20), ("GRC-002", 16), ("GRC-003", 30), ("GRC-004", 40), ("GRC-006", 48)],
            "Rice, flour, sugar, noodles",
        )
        self.add_purchase(
            "MAIN",
            "0112555003",
            "RECEIVED",
            dt(2026, 7, 22, 11, 5),
            [("DRY-001", 18), ("DRY-002", 24), ("DRY-003", 48), ("DRY-005", 12)],
            "Dairy cooler fill",
        )
        unilever_po = self.add_purchase(
            "MAIN",
            "0112555004",
            "RECEIVED",
            dt(2026, 7, 25, 14, 20),
            [("PER-001", 24), ("PER-002", 36), ("HLD-001", 18), ("HLD-002", 16), ("GRC-005", 20)],
            "Personal care and household",
        )
        self._unilever_po_main = unilever_po
        self.add_purchase(
            "MAIN",
            "0112555008",
            "RECEIVED",
            dt(2026, 8, 1, 7, 30),
            [("PRD-001", 20), ("PRD-002", 16), ("PRD-003", 30)],
            "Fresh produce — August 1 delivery",
        )
        self.add_purchase(
            "MAIN",
            "0112555005",
            "ORDERED",
            dt(2026, 8, 26, 16, 10),
            [("SNK-002", 40), ("SNK-003", 30), ("BKY-002", 20)],
            "Awaiting Maliban delivery",
        )
        self.add_purchase(
            "MAIN",
            "0112555006",
            "DRAFT",
            dt(2026, 8, 27, 15, 45),
            [("GRC-007", 12), ("GRC-008", 18), ("GRC-009", 16)],
            "Draft order — review prices",
        )
        self.add_purchase(
            "MAIN",
            "0112555007",
            "CANCELLED",
            dt(2026, 8, 10, 13, 0),
            [("FRZ-001", 10), ("FRZ-002", 8)],
            "Supplier could not deliver this week",
        )
        self.add_purchase(
            "MAIN",
            "0112555002",
            "PARTIAL",
            dt(2026, 8, 15, 10, 0),
            [("GRC-006", 24), ("GRC-010", 20)],
            "Partial shipment — remaining cartons pending",
        )

        self.add_purchase(
            "KDY",
            "0812200101",
            "RECEIVED",
            dt(2026, 7, 21, 11, 0),
            [("BEV-001", 24), ("BEV-004", 48), ("GRC-006", 24), ("SNK-002", 20)],
            "Opening Kandy grocery mix",
        )
        kdy_dairy = self.add_purchase(
            "KDY",
            "0812200103",
            "RECEIVED",
            dt(2026, 7, 24, 9, 20),
            [("DRY-001", 10), ("DRY-002", 12), ("DRY-003", 24)],
            "Dairy for Kandy cooler",
        )
        self._kdy_dairy_po = kdy_dairy
        self.add_purchase(
            "KDY",
            "0812200104",
            "RECEIVED",
            dt(2026, 8, 2, 8, 10),
            [("PRD-001", 12), ("PRD-003", 16)],
            "Kandy produce",
        )
        self.add_purchase(
            "KDY",
            "0812200102",
            "ORDERED",
            dt(2026, 8, 27, 12, 0),
            [("HLD-001", 10), ("PER-002", 16)],
            "Cleaning supplies on order",
        )

    def build_sales(self) -> None:
        main_customers = [
            "0771234501",
            "0771234502",
            "0771234503",
            "0771234504",
            "0771234505",
            "0771234506",
            "0771234507",
            "0771234508",
            "0771234509",
            "0771234511",
            "",
        ]
        main_cashiers = ["cashier@gmail.com", "nadee@gmail.com", "cashier@gmail.com"]
        payments = ["CASH", "CARD", "MOBILE", "CASH", "CREDIT", "CASH", "CARD"]
        baskets = [
            [("BEV-001", 2), ("SNK-001", 1)],
            [("BEV-004", 4), ("GRC-006", 2)],
            [("GRC-004", 1), ("GRC-010", 1), ("GRC-009", 1)],
            [("DRY-002", 2), ("BKY-001", 1)],
            [("SNK-002", 2), ("BEV-003", 1)],
            [("PER-002", 3), ("HLD-001", 1)],
            [("GRC-001", 1), ("GRC-005", 1)],
            [("PRD-001", 1), ("PRD-003", 2)],
            [("BEV-002", 1), ("SNK-003", 1), ("GRC-008", 2)],
            [("DRY-003", 6), ("BKY-002", 1)],
            [("HLD-003", 1), ("PER-001", 1)],
            [("FRZ-002", 1), ("BEV-005", 1)],
            [("GRC-006", 3), ("SNK-005", 1)],
            [("PRD-002", 1), ("GRC-007", 1)],
            [("BKY-001", 2), ("BEV-004", 2)],
            [("DRY-001", 1), ("GRC-003", 2)],
            [("HLD-002", 1), ("HLD-004", 1)],
            [("FRZ-001", 1), ("GRC-008", 1)],
            [("BEV-006", 1), ("SNK-004", 1)],
            [("BAB-001", 1), ("PER-003", 1)],
        ]

        sale_index = 0
        for day in range(1, 29):
            if day < 12:
                count = 1
            elif day < 28:
                count = 2
            else:
                count = 5
            for n in range(count):
                basket = baskets[sale_index % len(baskets)]
                if day == 28 and n == 4:
                    basket = [("BKY-003", 2)]  # sell remaining cake slices → out of stock
                hour = 9 + (n * 2) + (day % 3)
                minute = (sale_index * 7) % 60
                payment = payments[sale_index % len(payments)]
                customer = main_customers[sale_index % len(main_customers)]
                cashier = main_cashiers[sale_index % len(main_cashiers)]
                discount = 0
                if sale_index % 9 == 0:
                    discount = "0.50"
                elif sale_index % 11 == 0:
                    discount = "1.00"
                overpay = "1.00" if payment == "CASH" and sale_index % 5 == 0 else None
                created = dt(2026, 8, day, min(hour, 19), minute)
                ref = self.add_sale(
                    "MAIN",
                    created,
                    cashier,
                    customer or None,
                    basket,
                    payment_method=payment,
                    discount=discount,
                    overpay=overpay,
                )
                if day == 20 and n == 0:
                    self._main_return_sale = ref
                sale_index += 1

        kdy_baskets = [
            [("BEV-001", 1), ("SNK-002", 1)],
            [("BEV-004", 3), ("GRC-006", 1)],
            [("DRY-002", 1), ("BKY-001", 1)],
            [("PRD-001", 1), ("PRD-003", 1)],
            [("PER-002", 2), ("HLD-003", 1)],
            [("GRC-004", 1), ("GRC-010", 1)],
            [("SNK-001", 1), ("BEV-003", 1)],
            [("DRY-003", 4), ("BEV-004", 2)],
        ]
        kdy_customers = ["0752223301", "0752223302", "0752223303", "0752223305", ""]
        kdy_index = 0
        for day in range(8, 29, 2):
            count = 2 if day >= 20 else 1
            if day == 28:
                count = 3
            for n in range(count):
                basket = kdy_baskets[kdy_index % len(kdy_baskets)]
                hour = 10 + n * 3
                payment = ["CASH", "CARD", "MOBILE"][kdy_index % 3]
                created = dt(2026, 8, day, hour, 15)
                ref = self.add_sale(
                    "KDY",
                    created,
                    "user@gmail.com",
                    kdy_customers[kdy_index % len(kdy_customers)] or None,
                    basket,
                    payment_method=payment,
                    discount="0.50" if kdy_index % 4 == 0 else 0,
                    overpay="0.50" if payment == "CASH" else None,
                )
                if day == 22 and n == 0:
                    self._kdy_return_sale = ref
                kdy_index += 1

    def build_adjustments_and_returns(self) -> None:
        self.add_adjustment(
            "MAIN",
            "SNK-001",
            "DAMAGE",
            -3,
            "Crushed cartons found during shelf refill",
            "sudubole@gmail.com",
            dt(2026, 8, 12, 16, 20),
        )
        self.add_adjustment(
            "MAIN",
            "BEV-004",
            "FOUND",
            8,
            "Uncounted crate in the back store",
            "sudubole@gmail.com",
            dt(2026, 8, 14, 9, 10),
        )
        self.add_adjustment(
            "MAIN",
            "BKY-001",
            "LOSS",
            -2,
            "End-of-day stale bread written off",
            "nadee@gmail.com",
            dt(2026, 8, 18, 20, 5),
        )
        self.add_adjustment(
            "MAIN",
            "GRC-002",
            "CORRECTION",
            -1,
            "Cycle count shortfall on white rice",
            "sudubole@gmail.com",
            dt(2026, 8, 21, 11, 40),
        )
        self.add_adjustment(
            "KDY",
            "SNK-001",
            "DAMAGE",
            -1,
            "Torn packet on shop floor",
            "kandy.stock@popy.lk",
            dt(2026, 8, 19, 17, 0),
        )

        self.add_sales_return(
            "MAIN",
            self._main_return_sale,
            "Customer returned unwanted snacks — full refund",
            dt(2026, 8, 21, 11, 15),
        )
        self.add_sales_return(
            "KDY",
            self._kdy_return_sale,
            "Wrong item packed at checkout",
            dt(2026, 8, 23, 14, 40),
        )
        self.add_purchase_return(
            "MAIN",
            self._unilever_po_main,
            "Leaking dishwash bottles returned to supplier",
            dt(2026, 8, 8, 10, 30),
            items=[("HLD-001", 2)],
        )
        self.add_purchase_return(
            "KDY",
            self._kdy_dairy_po,
            "Short-dated yoghurt sent back",
            dt(2026, 8, 6, 15, 0),
            items=[("DRY-003", 6)],
        )

    def build(self) -> "SeedBuilder":
        self.build_catalog()
        self.build_purchases()
        self.build_sales()
        self.build_adjustments_and_returns()
        return self

    def stock_summary_rows(self) -> list[dict]:
        rows = []
        for (shop, sku), qty in sorted(self.stock.items()):
            product = self.products[(shop, sku)]
            reorder = product["reorder_level"]
            if qty <= 0:
                status = "out"
            elif qty <= reorder:
                status = "low"
            else:
                status = "in"
            rows.append(
                {
                    "shop_code": shop,
                    "sku": sku,
                    "name": product["name"],
                    "opening_stock": product["opening_stock"],
                    "final_stock": qty,
                    "reorder_level": reorder,
                    "stock_status": status,
                }
            )
        return rows


def customer_rows() -> list[dict]:
    rows = []
    for shop_code, people in CUSTOMERS.items():
        for name, phone, email, address, points in people:
            rows.append(
                {
                    "shop_code": shop_code,
                    "name": name,
                    "phone": phone,
                    "email": email or "",
                    "address": address,
                    "loyalty_points": points,
                }
            )
    return rows


def supplier_rows() -> list[dict]:
    rows = []
    for shop_code, people in SUPPLIERS.items():
        for name, company, phone, email, address in people:
            rows.append(
                {
                    "shop_code": shop_code,
                    "name": name,
                    "company_name": company,
                    "phone": phone,
                    "email": email,
                    "address": address,
                }
            )
    return rows


def notification_rows() -> list[dict]:
    return [
        {
            "shop_code": "MAIN",
            "pos_checkout_email_enabled": True,
            "pos_checkout_sms_enabled": False,
            "pos_checkout_cashier_email_enabled": True,
            "pos_checkout_cashier_sms_enabled": False,
            "low_inventory_email_enabled": True,
            "low_inventory_sms_enabled": False,
            "low_inventory_alert_phone": "0771234599",
            "new_customer_email_enabled": True,
            "new_customer_sms_enabled": False,
            "new_user_email_enabled": True,
            "new_user_sms_enabled": False,
        },
        {
            "shop_code": "KDY",
            "pos_checkout_email_enabled": True,
            "pos_checkout_sms_enabled": False,
            "pos_checkout_cashier_email_enabled": True,
            "pos_checkout_cashier_sms_enabled": False,
            "low_inventory_email_enabled": True,
            "low_inventory_sms_enabled": False,
            "low_inventory_alert_phone": "0752223399",
            "new_customer_email_enabled": True,
            "new_customer_sms_enabled": False,
            "new_user_email_enabled": True,
            "new_user_sms_enabled": False,
        },
    ]


HEADER_FILL = PatternFill("solid", fgColor="1B5E20")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="E8F5E9")
THIN = Border(
    left=Side(style="thin", color="C8E6C9"),
    right=Side(style="thin", color="C8E6C9"),
    top=Side(style="thin", color="C8E6C9"),
    bottom=Side(style="thin", color="C8E6C9"),
)


def write_xlsx(path: Path, headers: list[str], rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = path.stem[:31]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, row in enumerate(rows, start=2):
        values = []
        for key in headers:
            value = row.get(key, "")
            if isinstance(value, Decimal):
                value = float(value)
            elif isinstance(value, bool):
                value = "TRUE" if value else "FALSE"
            elif value is None:
                value = ""
            values.append(value)
        ws.append(values)
        fill = ALT_FILL if i % 2 == 0 else None
        for cell in ws[i]:
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in rows[:80]:
            max_len = max(max_len, len(str(row.get(header, "") or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 3), 42)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def seed_files(builder: SeedBuilder) -> list[tuple[str, list[str], list[dict]]]:
    return [
        ("shops.xlsx", ["code", "name", "address", "phone", "email", "is_active"], SHOPS),
        (
            "users.xlsx",
            ["email", "password", "name", "role", "shop_code", "is_active", "is_staff", "is_superuser"],
            USERS,
        ),
        (
            "notification_settings.xlsx",
            [
                "shop_code",
                "pos_checkout_email_enabled",
                "pos_checkout_sms_enabled",
                "pos_checkout_cashier_email_enabled",
                "pos_checkout_cashier_sms_enabled",
                "low_inventory_email_enabled",
                "low_inventory_sms_enabled",
                "low_inventory_alert_phone",
                "new_customer_email_enabled",
                "new_customer_sms_enabled",
                "new_user_email_enabled",
                "new_user_sms_enabled",
            ],
            notification_rows(),
        ),
        ("categories.xlsx", ["shop_code", "name", "description"], builder.category_rows),
        (
            "products.xlsx",
            [
                "shop_code",
                "name",
                "sku",
                "barcode",
                "category_name",
                "brand",
                "unit",
                "cost_price",
                "selling_price",
                "reorder_level",
                "opening_stock",
                "status",
            ],
            builder.product_rows,
        ),
        (
            "customers.xlsx",
            ["shop_code", "name", "phone", "email", "address", "loyalty_points"],
            customer_rows(),
        ),
        (
            "suppliers.xlsx",
            ["shop_code", "name", "company_name", "phone", "email", "address"],
            supplier_rows(),
        ),
        (
            "purchases.xlsx",
            ["shop_code", "reference", "supplier_phone", "status", "note", "created_at"],
            builder.purchase_rows,
        ),
        (
            "purchase_items.xlsx",
            ["shop_code", "purchase_reference", "sku", "quantity", "cost_price"],
            builder.purchase_item_rows,
        ),
        (
            "sales.xlsx",
            [
                "shop_code",
                "reference",
                "customer_phone",
                "cashier_email",
                "payment_method",
                "discount",
                "tax",
                "subtotal",
                "total",
                "amount_paid",
                "created_at",
            ],
            builder.sale_rows,
        ),
        (
            "sale_items.xlsx",
            ["shop_code", "sale_reference", "sku", "quantity", "unit_price"],
            builder.sale_item_rows,
        ),
        (
            "stock_adjustments.xlsx",
            ["shop_code", "sku", "adjustment_type", "quantity", "note", "user_email", "created_at"],
            builder.adjustment_rows,
        ),
        (
            "sales_returns.xlsx",
            ["shop_code", "reference", "sale_reference", "reason", "refund_amount", "items", "created_at"],
            builder.sales_return_rows,
        ),
        (
            "purchase_returns.xlsx",
            ["shop_code", "reference", "purchase_reference", "reason", "amount", "items", "created_at"],
            builder.purchase_return_rows,
        ),
    ]


def write_seed_files(out_dir: Path) -> SeedBuilder:
    builder = SeedBuilder().build()
    out_dir.mkdir(parents=True, exist_ok=True)
    for name, headers, rows in seed_files(builder):
        write_xlsx(out_dir / name, headers, rows)
    return builder


def main() -> None:
    out_dir = Path(__file__).resolve().parent
    combined = out_dir / "popy_example_data.xlsx"
    if combined.exists():
        combined.unlink()
    builder = write_seed_files(out_dir)
    files = seed_files(builder)
    summary = builder.stock_summary_rows()
    low_n = sum(1 for row in summary if row["stock_status"] == "low")
    out_n = sum(1 for row in summary if row["stock_status"] == "out")
    print(f"Wrote {len(files)} xlsx files in {out_dir}")
    for name, _, rows in files:
        print(f"  {name}: {len(rows)} rows")
    print(f"  expected stock: in={len(summary) - low_n - out_n} low={low_n} out={out_n}")


if __name__ == "__main__":
    main()
